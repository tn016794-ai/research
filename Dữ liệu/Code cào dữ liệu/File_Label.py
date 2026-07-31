"""
Gắn nhãn (label) rủi ro lỗi thời cho các đời iPhone.

Nguồn dữ liệu:
  1. Wikipedia - "List of iPhone models"      -> ngày ngừng bán (Discontinued)
  2. The Apple Wiki - "List of iPhones"       -> phiên bản iOS cuối cùng máy được cập nhật (Last/Latest firmware)
  3. setapp - "The full list of iOS versions" -> ngày phát hành từng phiên bản iOS lớn (major)

Công thức:
  hardware_lifespan = ngày ngừng bán (Discontinued)      + 5 năm
  software_lifespan = ngày phát hành iOS lớn tiếp theo mà máy KHÔNG được cập nhật + 3 năm
  real_lifecycle     = MIN(hardware_lifespan, software_lifespan)   (nguyên lý "ống gỗ ngắn nhất" / Cannikin Law:
                        tuổi thọ thực tế của thiết bị bị giới hạn bởi yếu tố "ngắn" hơn, phần cứng hay phần mềm)
"""
import re
import datetime as dt
from dateutil.relativedelta import relativedelta
from bs4 import BeautifulSoup
from grid_parser import table_to_grid
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Đã sửa lại toàn bộ đường dẫn đầu vào
WIKI_FILE = 'Data cho NCKH/List of iPhone models (Label_1)- Wikipedia.html'
APPLEWIKI_FILE = 'Data cho NCKH/List of iPhones - The Apple Wiki.html'
IOS_FILE = 'Data cho NCKH/The full list of iOS versions_ From iOS 1 to latest iOS 26 version.html'
DATE_TAG_RE = re.compile(r'([A-Z][a-z]+ \d{1,2}, \d{4}) \(\s*(\d{4}-\d{2}-\d{2})\s*\)(?:\s*\(([^)]+)\))?')

# ============================================================
# 1. Wikipedia: bảng "List of iPhone models" -> ngày ngừng bán
# ============================================================
with open(WIKI_FILE, encoding='utf-8') as f:
    soup = BeautifulSoup(f, 'lxml')
wiki_table = soup.find_all('table')[0]
grid = table_to_grid(wiki_table)
device_rows = grid[2:-3]  # bỏ 2 dòng header + 3 dòng chú thích/legend cuối bảng


def split_family(name):
    """'iPhone 12 / 12 mini' -> ['iPhone 12', 'iPhone 12 mini']"""
    parts = [p.strip() for p in name.split(' / ')]
    if len(parts) == 1:
        return parts
    out = [parts[0]]
    for p in parts[1:]:
        out.append('iPhone ' + p)
    return out


def extract_date_tags(raw):
    return DATE_TAG_RE.findall(raw)


families = []
for row in device_rows:
    family_name = row[0]
    discontinued_raw = row[4]
    families.append({
        'family_name': family_name,
        'models': split_family(family_name),
        'discontinued_tags': extract_date_tags(discontinued_raw),
        'discontinued_raw': discontinued_raw,
    })


def assign_discontinued(fam):
    """Gán ngày ngừng bán cho từng model riêng lẻ trong 1 'family' (vd iPhone 6 / 6 Plus)."""
    models = fam['models']
    tags = fam['discontinued_tags']
    result = {}
    if not tags:
        for m in models:
            result[m] = None          # vẫn đang được bán / không rõ
        return result
    if len(tags) == 1:
        iso = tags[0][1]
        for m in models:
            result[m] = iso
        return result
    tag_texts = [tg.strip().lower() for _, _, tg in tags]
    is_region_only = all(('eea' in tt or tt == '') for tt in tag_texts)
    if is_region_only:
        # Chênh lệch chỉ do khu vực (EEA vs ngoài EEA) -> lấy ngày "Outside EEA" làm mốc chung
        chosen = None
        for _, iso, tg in tags:
            if 'outside' in tg.lower():
                chosen = iso
        if chosen is None:
            chosen = max(iso for _, iso, _ in tags)
        for m in models:
            result[m] = chosen
        return result
    untagged_iso = None
    for _, iso, tg in tags:
        if tg.strip() == '':
            untagged_iso = iso
    for m in models:
        suffix = re.sub(r'^iPhone\s+', '', m).strip().lower()
        best = None
        for _, iso, tg in tags:
            tgl = tg.strip().lower()
            if tgl and tgl == suffix:
                best = iso
                break
        if best is None:
            best = untagged_iso if untagged_iso is not None else max(iso for _, iso, _ in tags)
        result[m] = best
    return result


model_discontinued = {}
for fam in families:
    model_discontinued.update(assign_discontinued(fam))

# ============================================================
# 2. The Apple Wiki: phiên bản iOS cuối cùng mỗi máy được cập nhật
# ============================================================
with open(APPLEWIKI_FILE, encoding='utf-8') as f:
    soup2 = BeautifulSoup(f, 'lxml')
content = soup2.find('div', class_='mw-parser-output')
top_section = content.find('section')
device_sections = top_section.find_all('section', recursive=False)

FW_LAST_RE = re.compile(
    r'Last firmware:(.*?)(?:Latest beta firmware|Internal Name|Firmware Identifiers|RAM:|Storage:|Wi-Fi:|Bluetooth|$)')
FW_LATEST_RE = re.compile(
    r'Latest (?:public )?firmware:(.*?)(?:Latest beta firmware|Internal Name|Firmware Identifiers|RAM:|Storage:|Wi-Fi:|Bluetooth|$)')

apple_wiki_data = {}
for s in device_sections:
    h = s.find(['h1', 'h2', 'h3'])
    if not h:
        continue
    name = h.get_text(' ', strip=True)
    if name in ('Notes', 'References'):
        continue
    txt = s.get_text(' ', strip=True)
    m_last = FW_LAST_RE.search(txt)
    m_latest = FW_LATEST_RE.search(txt)
    if m_last:
        fw_text, status = m_last.group(1).strip(), 'ended'       # máy đã ngừng nhận cập nhật iOS
    elif m_latest:
        fw_text, status = m_latest.group(1).strip(), 'ongoing'   # máy vẫn đang được hỗ trợ
    else:
        fw_text, status = None, 'unknown'
    version = None
    if fw_text:
        vm = re.search(r'(\d+(?:\.\d+)*)', fw_text)
        if vm:
            version = vm.group(1)
    apple_wiki_data[name] = {'fw_text': fw_text, 'status': status, 'version': version}

# Chuẩn hoá tên gọi khác nhau giữa 2 nguồn
NAME_ALIAS = {
    'iPhone 4s': 'iPhone 4S',
    'iPhone SE (1st)': 'iPhone SE (1st generation)',
    'iPhone SE (2nd)': 'iPhone SE (2nd generation)',
    'iPhone SE (3rd)': 'iPhone SE (3rd generation)',
}

# ============================================================
# 3. setapp: ngày phát hành từng phiên bản iOS lớn (major)
# ============================================================
with open(IOS_FILE, encoding='utf-8') as f:
    soup3 = BeautifulSoup(f, 'lxml')
ios_table = soup3.find_all('table')[0]
ios_rows = ios_table.find_all('tr')

ios_release = {}  # major (int) -> datetime.date
for r in ios_rows[1:]:
    cells = r.find_all(['th', 'td'])
    if len(cells) < 2:
        continue
    ver_text = cells[0].get_text(' ', strip=True)
    date_text = cells[1].get_text(' ', strip=True)
    vm = re.search(r'iOS\s+(\d+)', ver_text)
    if not vm:
        continue
    major = int(vm.group(1))
    try:
        d = dt.datetime.strptime(date_text, '%B %d, %Y').date()
    except ValueError:
        continue  # vd "Upcoming in September, 2026" cho iOS 27 -> chưa có ngày chính thức, bỏ qua
    ios_release[major] = d

ordered_majors = sorted(ios_release.keys())  # [1,2,...,18,26]


def next_major_release(last_major):
    """Trả về ngày phát hành phiên bản iOS lớn kế tiếp mà máy KHÔNG còn được cập nhật."""
    idx = None
    for i, m in enumerate(ordered_majors):
        if m >= last_major:
            idx = i
            break
    if idx is None:
        return None
    if last_major == ordered_majors[idx]:
        idx += 1
    if idx >= len(ordered_majors):
        return None
    return ios_release[ordered_majors[idx]]


def parse_iso_date(s):
    return dt.date.fromisoformat(s)


# ============================================================
# 4. Tổng hợp & tính toán cho từng model
# ============================================================
rows_out = []
for model in model_discontinued:
    aw_key = NAME_ALIAS.get(model, model)
    aw = apple_wiki_data.get(aw_key)

    # ---- Hardware ----
    disc_iso = model_discontinued[model]
    hw_discontinued_date = parse_iso_date(disc_iso) if disc_iso else None
    hw_lifespan_date = hw_discontinued_date + relativedelta(years=5) if hw_discontinued_date else None

    # ---- Software ----
    sw_last_version = aw['version'] if aw else None
    sw_status = aw['status'] if aw else 'unknown'
    sw_lifespan_date = None
    next_ios_date = None
    if aw and sw_status == 'ended' and sw_last_version:
        last_major = int(sw_last_version.split('.')[0])
        next_ios_date = next_major_release(last_major)
        if next_ios_date:
            sw_lifespan_date = next_ios_date + relativedelta(years=3)

    # ---- Real lifecycle = MIN (nguyên lý "ống gỗ ngắn nhất") ----
    candidates = [d for d in (hw_lifespan_date, sw_lifespan_date) if d is not None]
    if candidates:
        real_lifecycle_date = min(candidates)
        limiting_factor = 'Phần cứng' if (sw_lifespan_date is None or hw_lifespan_date <= sw_lifespan_date) else 'Phần mềm'
    else:
        real_lifecycle_date = None
        limiting_factor = None

    rows_out.append({
        'Device Name': model,
        'Discontinued Date': hw_discontinued_date,
        'Hardware Lifespan (Discontinued + 5y)': hw_lifespan_date,
        'Last Supported iOS': sw_last_version,
        'Software Support Status': 'Đã dừng cập nhật' if sw_status == 'ended' else ('Đang được hỗ trợ' if sw_status == 'ongoing' else 'Không rõ'),
        'Next iOS Not Supported': next_ios_date,
        'Software Lifespan (Next iOS + 3y)': sw_lifespan_date,
        'Real Lifecycle (MIN rule)': real_lifecycle_date,
        'Limiting Factor': limiting_factor,
    })


# ============================================================
# 5. Xuất ra file Excel
# ============================================================
wb = Workbook()
ws = wb.active
ws.title = "iPhone Obsolescence Labels"

FONT_NAME = "Arial"
header_font = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="1F4E78")
normal_font = Font(name=FONT_NAME, size=10)
bold_font = Font(name=FONT_NAME, size=10, bold=True)
note_font = Font(name=FONT_NAME, size=9, italic=True, color="666666")
thin = Side(style="thin", color="B7B7B7")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
hw_fill = PatternFill("solid", fgColor="FCE4D6")   # cam nhạt - phần cứng là yếu tố giới hạn
sw_fill = PatternFill("solid", fgColor="D9E1F2")   # xanh nhạt - phần mềm là yếu tố giới hạn
active_fill = PatternFill("solid", fgColor="E2EFDA")  # xanh lá nhạt - đang hoạt động

headers = [
    "Device Name",
    "Discontinued Date\n(nguồn: Wikipedia)",
    "Hardware Lifespan\n(Discontinued + 5 năm)",
    "Last Supported iOS\n(nguồn: The Apple Wiki)",
    "Software Support Status",
    "Next iOS Not Supported\n(nguồn: setapp - ngày phát hành)",
    "Software Lifespan\n(Next iOS release + 3 năm)",
    "Real Lifecycle\n(MIN rule - nguyên lý ống gỗ ngắn nhất)",
    "Limiting Factor\n(Yếu tố giới hạn)",
]

ws.append(headers)
for col_idx in range(1, len(headers) + 1):
    c = ws.cell(row=1, column=col_idx)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border
ws.row_dimensions[1].height = 45

DATE_FMT = "yyyy-mm-dd"

def fmt_date_cell(cell, value):
    if isinstance(value, dt.date):
        cell.value = value
        cell.number_format = DATE_FMT
    else:
        cell.value = value if value else "—"

r = 2
for row in rows_out:
    ws.cell(row=r, column=1, value=row['Device Name'])
    fmt_date_cell(ws.cell(row=r, column=2), row['Discontinued Date'])
    fmt_date_cell(ws.cell(row=r, column=3), row['Hardware Lifespan (Discontinued + 5y)'])
    ws.cell(row=r, column=4, value=row['Last Supported iOS'] or "—")
    ws.cell(row=r, column=5, value=row['Software Support Status'])
    fmt_date_cell(ws.cell(row=r, column=6), row['Next iOS Not Supported'])
    fmt_date_cell(ws.cell(row=r, column=7), row['Software Lifespan (Next iOS + 3y)'])
    fmt_date_cell(ws.cell(row=r, column=8), row['Real Lifecycle (MIN rule)'])
    ws.cell(row=r, column=9, value=row['Limiting Factor'] or "Đang hoạt động (chưa xác định)")

    fill = None
    if row['Limiting Factor'] == 'Phần cứng':
        fill = hw_fill
    elif row['Limiting Factor'] == 'Phần mềm':
        fill = sw_fill
    else:
        fill = active_fill

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=r, column=col_idx)
        cell.font = normal_font
        cell.border = border
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if col_idx == 1:
            cell.font = bold_font
            cell.alignment = Alignment(horizontal="left", vertical="center")
    r += 1

# Column widths
widths = [22, 16, 16, 16, 16, 18, 16, 18, 16]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "A2"

# ---- Notes sheet ----
notes = wb.create_sheet("Ghi chú & Nguồn")
notes_content = [
    ("PHƯƠNG PHÁP TÍNH TOÁN", True),
    ("1. Hardware Lifespan = Ngày ngừng bán (Discontinued) + 5 năm.", False),
    ("   Nguồn: Wikipedia - 'List of iPhone models' (cột Discontinued).", False),
    ("2. Software Lifespan = Ngày phát hành phiên bản iOS lớn (major) TIẾP THEO mà máy", False),
    ("   KHÔNG còn được cập nhật + 3 năm.", False),
    ("   - Phiên bản iOS cuối cùng máy nhận được: nguồn The Apple Wiki (mục 'Last firmware'/'Latest firmware').", False),
    ("   - Ngày phát hành từng phiên bản iOS lớn: nguồn setapp - 'The full list of iOS versions'.", False),
    ("3. Real Lifecycle = MIN(Hardware Lifespan, Software Lifespan)", False),
    ("   Áp dụng nguyên lý 'ống gỗ ngắn nhất' (Cannikin Law / short-plank principle): tuổi thọ", False),
    ("   thực tế của thiết bị bị giới hạn bởi yếu tố tới hạn SỚM HƠN giữa phần cứng và phần mềm.", False),
    ("", False),
    ("QUY ƯỚC MÀU", True),
    ("Cam nhạt: yếu tố giới hạn là PHẦN CỨNG (hardware lifespan <= software lifespan)", False),
    ("Xanh dương nhạt: yếu tố giới hạn là PHẦN MỀM (software lifespan < hardware lifespan)", False),
    ("Xanh lá nhạt: máy vẫn đang được bán và/hoặc vẫn đang nhận cập nhật iOS -> chưa xác định real lifecycle", False),
    ("", False),
    ("GIẢ ĐỊNH & LƯU Ý QUAN TRỌNG (đọc trước khi dùng cho NCKH)", True),
    ("- Với các dòng máy Wikipedia gộp chung 2 model (vd 'iPhone 6 / 6 Plus'), ngày ngừng bán được", False),
    ("  tách theo từng model dựa trên chú thích trong bảng gốc (vd '(12 mini)', '(32 GB model)').", False),
    ("  Một số biến thể dung lượng lưu trữ (vd bản 32GB của 6/6 Plus) được BỎ QUA, chỉ lấy ngày ngừng", False),
    ("  bán của biến thể chính — có thể lệch vài tháng so với thực tế với biến thể phụ.", False),
    ("- Với iPhone SE (3rd)/14/14 Plus có 2 ngày ngừng bán khác nhau theo khu vực (EEA vs ngoài EEA),", False),
    ("  script lấy ngày 'Outside EEA' (ngày ngừng bán ở thị trường ngoài châu Âu) làm mốc chung.", False),
    ("- Với iPhone 4 (có 2 phiên bản GSM/CDMA release date khác nhau), chỉ lấy 1 ngày ngừng bán chung", False),
    ("  theo bảng Wikipedia (bảng gốc không tách riêng ngày ngừng bán CDMA).", False),
    ("- 'Next iOS Not Supported' được xác định theo TÊN phiên bản major, có xử lý việc Apple đổi cách", False),
    ("  đặt tên phiên bản (nhảy từ iOS 18 thẳng lên iOS 26 vào năm 2025), không có iOS 19-25.", False),
    ("- Các máy đang hiển thị 'Đang được hỗ trợ' (status = Latest firmware) chưa có Software Lifespan vì", False),
    ("  Apple chưa dừng cập nhật — Real Lifecycle của các máy này tạm lấy theo Hardware Lifespan và sẽ", False),
    ("  cần cập nhật lại khi Apple chính thức dừng hỗ trợ phần mềm cho model đó.", False),
    ("- iPhone 16/16 Plus, 17, 17 Pro/Pro Max, 17e, iPhone Air: tại thời điểm lấy dữ liệu (07/2026) vẫn", False),
    ("  đang được bán ('In production') nên chưa có Hardware Lifespan -> Real Lifecycle để trống.", False),
    ("", False),
    ("NGUỒN DỮ LIỆU GỐC", True),
    ("1. Wikipedia — List of iPhone models", False),
    ("2. The Apple Wiki — List of iPhones", False),
    ("3. setapp.com — The full list of iOS versions (From iOS 1 to latest iOS 26 version)", False),
    (f"Ngày tổng hợp dữ liệu / xử lý: {dt.date.today().isoformat()}", False),
]
nr = 1
for text, is_header in notes_content:
    c = notes.cell(row=nr, column=1, value=text)
    c.font = Font(name=FONT_NAME, size=11, bold=True) if is_header else Font(name=FONT_NAME, size=10)
    nr += 1
notes.column_dimensions['A'].width = 100

# Đã sửa lại đường dẫn đầu ra
wb.save('Data cho NCKH/iPhone_Obsolescence_Labels.xlsx')
print("Đã lưu file thành công!")
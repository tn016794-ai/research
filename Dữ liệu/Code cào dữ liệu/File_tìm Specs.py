"""
Trích xuất 4 thông số kỹ thuật từ The Apple Wiki (List of iPhones) và quy đổi về dạng số (numerical)
để phục vụ mô hình dự báo rủi ro lỗi thời iPhone.

Quy đổi:
  - Bluetooth   : lấy trực tiếp số phiên bản (vd "Bluetooth 5.3" -> 5.3)
  - Camera      : lấy số Megapixel của camera CHÍNH (rear/main) (vd "Rear: 48MP main..." -> 48.0)
  - Cellular    : lấy số thế hệ mạng (G) CAO NHẤT thiết bị hỗ trợ (vd "LTE (4G)", "5G (sub-6GHz)" -> 4 hoặc 5)
  - Wi-Fi       : quy về thang 3-7 (802.11b/g=3, b/g/n=4, ac=5, ax/Wi-Fi 6=6, be/Wi-Fi 7=7)

Input : //Users/macbook/Library/CloudStorage/OneDrive-ut.edu.vn/Data cho NCKH/List of iPhones - The Apple Wiki.html
Output: //Users/macbook/Library/CloudStorage/OneDrive-ut.edu.vn/Data cho NCKH/iPhone_Connectivity_Specs.xlsx
"""
import re
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

FILE = '//Users/macbook/Library/CloudStorage/OneDrive-ut.edu.vn/Data cho NCKH/List of iPhones - The Apple Wiki.html'
OUT_FILE = '//Users/macbook/Library/CloudStorage/OneDrive-ut.edu.vn/Data cho NCKH/iPhone_Connectivity_Specs.xlsx'

# ============================================================
# 1. Parse từng section thiết bị trong trang The Apple Wiki
# ============================================================
with open(FILE, encoding='utf-8') as f:
    soup = BeautifulSoup(f, 'lxml')
content = soup.find('div', class_='mw-parser-output')
top_section = content.find('section')
device_sections = top_section.find_all('section', recursive=False)

BT_RE = re.compile(r'Bluetooth\s+(\d+(?:\.\d+)?)')
REAR_RE = re.compile(r'Rear:\s*(?:\d+\s*[×xX]\s*)?(\d+(?:\.\d+)?)\s*(?:MP|megapixels)', re.IGNORECASE)
CELL_G_RE = re.compile(r'(\d+(?:\.\d+)?)\s*G\b')
WIFI_DIRECT_RE = re.compile(r'Wi.?Fi\s*(\d)E?\b')      # bắt "Wi-Fi 6", "Wi-Fi 6E", "Wi-Fi 7", "(Wi-Fi 6)"
WIFI_CODE_RE = re.compile(r'802\.11([a-z/]+)')          # fallback: suy ra từ chuẩn 802.11x

results = []
for s in device_sections:
    h = s.find(['h1', 'h2', 'h3'])
    if not h:
        continue
    name = h.get_text(' ', strip=True)
    if name in ('Notes', 'References'):
        continue
    txt = s.get_text(' ', strip=True)

    # ---- Bluetooth: lấy số phiên bản đầu tiên ----
    m = BT_RE.search(txt)
    bluetooth = float(m.group(1)) if m else None

    # ---- Camera: Megapixel camera chính (rear/main) ----
    m = REAR_RE.search(txt)
    camera_mp = float(m.group(1)) if m else None

    # ---- Cellular: thế hệ mạng (G) cao nhất ----
    matches = CELL_G_RE.findall(txt)
    if matches:
        cellular_g = max(float(x) for x in matches)
    elif 'LTE' in txt:
        cellular_g = 4.0   # trang không ghi rõ số G nhưng có "Gigabit-class LTE" -> quy ước 4G
    elif 'EDGE' in txt:
        cellular_g = 2.5
    else:
        cellular_g = None

    # ---- Wi-Fi: quy về thang 3-7 ----
    m = WIFI_DIRECT_RE.search(txt)
    if m:
        wifi_gen = int(m.group(1))
    else:
        wifi_gen = None
        m2 = WIFI_CODE_RE.search(txt)
        if m2:
            codes = m2.group(1).split('/')
            if 'be' in codes:
                wifi_gen = 7
            elif 'ax' in codes:
                wifi_gen = 6
            elif 'ac' in codes:
                wifi_gen = 5
            elif 'n' in codes:
                wifi_gen = 4
            elif 'g' in codes:
                wifi_gen = 3

    results.append({
        'Device Name': name,
        'Bluetooth': bluetooth,
        'Camera (MP)': camera_mp,
        'Cellular (G)': cellular_g,
        'Wi-Fi (Gen)': wifi_gen,
    })

print(f'Đã trích xuất {len(results)} thiết bị.')

# ============================================================
# 2. Xuất ra Excel
# ============================================================
wb = Workbook()
ws = wb.active
ws.title = "Connectivity Specs"

FONT_NAME = "Arial"
header_font = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="1F4E78")
normal_font = Font(name=FONT_NAME, size=10)
bold_font = Font(name=FONT_NAME, size=10, bold=True)
thin = Side(style="thin", color="B7B7B7")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

headers = ['Device Name', 'Bluetooth', 'Camera (MP)', 'Cellular (G)', 'Wi-Fi (Gen)']
ws.append(headers)
for col_idx in range(1, len(headers) + 1):
    c = ws.cell(row=1, column=col_idx)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border

r = 2
for row in results:
    ws.cell(row=r, column=1, value=row['Device Name'])
    ws.cell(row=r, column=2, value=row['Bluetooth'])
    ws.cell(row=r, column=3, value=row['Camera (MP)'])
    ws.cell(row=r, column=4, value=row['Cellular (G)'])
    ws.cell(row=r, column=5, value=row['Wi-Fi (Gen)'])
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=r, column=col_idx)
        cell.font = bold_font if col_idx == 1 else normal_font
        cell.border = border
        cell.alignment = Alignment(horizontal="left" if col_idx == 1 else "center", vertical="center")
    r += 1

widths = [26, 12, 14, 14, 14]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"

# ---- Sheet ghi chú ----
notes = wb.create_sheet("Ghi chú quy đổi")
notes_content = [
    ("QUY TẮC QUY ĐỔI SANG SỐ (NUMERICAL)", True),
    ("1. Bluetooth: lấy trực tiếp số phiên bản ghi trên trang, vd 'Bluetooth 5.3' -> 5.3.", False),
    ("2. Camera: lấy số Megapixel của camera CHÍNH (rear/main), bỏ qua ống kính phụ (ultrawide/telephoto).", False),
    ("   Vd 'Rear: 48MP main; 12MP ultrawide...' -> lấy 48.0.", False),
    ("3. Cellular: lấy thế hệ mạng (G) CAO NHẤT được đề cập trong trang cho thiết bị đó.", False),
    ("   Vd có cả '(3G)' và '(4G)' -> lấy 4.0. Có '5G' -> lấy 5.0.", False),
    ("4. Wi-Fi: quy về thang số nguyên 3-7 (iPhone đầu tiên = Wi-Fi 3, tương ứng 802.11b/g):", False),
    ("     802.11b/g       -> 3", False),
    ("     802.11b/g/n hoặc a/b/g/n -> 4", False),
    ("     802.11ac (Wi-Fi 5) -> 5", False),
    ("     802.11ax / Wi-Fi 6 / Wi-Fi 6E -> 6", False),
    ("     802.11be / Wi-Fi 7 -> 7", False),
    ("", False),
    ("GIẢ ĐỊNH / LƯU Ý QUAN TRỌNG", True),
    ("- iPhone SE (2nd generation) và iPhone SE (3rd generation): trang The Apple Wiki không ghi rõ", False),
    ("  số G cho mạng di động (chỉ ghi 'Gigabit-class LTE'), nên quy ước gán Cellular = 4.0 (4G).", False),
    ("  Lưu ý: trên thực tế iPhone SE (3rd generation) CÓ hỗ trợ 5G, nhưng nguồn dữ liệu này không", False),
    ("  ghi rõ nên script lấy theo đúng thông tin có trong trang -> nên kiểm tra thủ công nếu cần độ", False),
    ("  chính xác tuyệt đối cho model này.", False),
    ("- Wi-Fi 6E được quy về cùng mức 6 với Wi-Fi 6 (không tách biệt do thang đo user yêu cầu là số", False),
    ("  nguyên 3-7).", False),
    ("", False),
    ("NGUỒN DỮ LIỆU", True),
    ("The Apple Wiki - List of iPhones (https://theapplewiki.com)", False),
]
nr = 1
for text, is_header in notes_content:
    c = notes.cell(row=nr, column=1, value=text)
    c.font = Font(name=FONT_NAME, size=11, bold=True) if is_header else Font(name=FONT_NAME, size=10)
    nr += 1
notes.column_dimensions['A'].width = 100

wb.save(OUT_FILE)
print('Đã lưu:', OUT_FILE)
